from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from openpyxl.utils import coordinate_to_tuple

import os
import win32com.client as win32


# -------------------------
# 🔧 UTILIDADES
# -------------------------

def formatear_fecha(fecha_str):
    try:
        return datetime.strptime(fecha_str, "%d.%m.%Y")
    except:
        return fecha_str


def auto_alto_fila(ws, fila, texto, ancho_aprox=40):
    if not texto:
        ws.row_dimensions[fila].height = 20
        return

    lineas = len(str(texto)) // ancho_aprox + 1
    ws.row_dimensions[fila].height = max(20, lineas * 15)


def limpiar_tabla(ws, fila_inicio, max_filas=17):
    for fila in range(fila_inicio, fila_inicio + max_filas):
        for col in ["D", "E"]:
            celda = f"{col}{fila}"
            celda_real = get_celda_principal(ws, celda)

            fila_real, _ = coordinate_to_tuple(celda_real)

            if fila_real >= fila_inicio:
                ws[celda_real].value = None


# -------------------------
# 🧠 MANEJO DE MERGE
# -------------------------

def get_celda_principal(ws, celda):
    fila, col = coordinate_to_tuple(celda)

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        if min_row <= fila <= max_row and min_col <= col <= max_col:
            return merged_range.start_cell.coordinate

    return celda


def set_valor(ws, celda, valor):
    celda_real = get_celda_principal(ws, celda)
    ws[celda_real] = valor
    return celda_real


def set_wrap(ws, celda):
    celda_real = get_celda_principal(ws, celda)
    ws[celda_real].alignment = Alignment(wrap_text=True)
    return celda_real


# -------------------------
# 🔥 MARCA DE AGUA (PYWIN32)
# -------------------------

def agregar_marca_agua_pywin32(ruta_excel):
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(os.path.abspath(ruta_excel))
    ws = wb.ActiveSheet

    try:
        shape = ws.Shapes.AddPicture(
            Filename=os.path.abspath("plantillas/marca_agua.png"),
            LinkToFile=False,
            SaveWithDocument=True,
            Left=80,     # columna B aprox
            Top=350,     # fila 39 aprox
            Width=600,
            Height=350
        )

        # enviar al fondo
        shape.ZOrder(1)

        # transparencia
        shape.Fill.Transparency = 0.7

    except Exception as e:
        print("Error watermark:", e)

    wb.Save()
    wb.Close()
    excel.Quit()


# -------------------------
# 🚀 GENERADOR PRINCIPAL
# -------------------------

def llenar_excel(ruta_plantilla, ruta_salida, datos):
    wb = load_workbook(ruta_plantilla)
    ws = wb.active

    # -------------------------
    # LIMPIAR TABLA
    # -------------------------
    fila_inicio = 37
    limpiar_tabla(ws, fila_inicio)

    # -------------------------
    # DATOS FIJOS
    # -------------------------
    set_valor(ws, "E12", datos.get("bl", ""))

    set_valor(ws, "B17", datos.get("consignee", ""))
    set_valor(ws, "B18", datos.get("nit", ""))

    set_valor(ws, "B23", datos.get("consignee", ""))
    set_valor(ws, "B24", datos.get("nit", ""))

    set_valor(ws, "B32", datos.get("nave", ""))
    set_valor(ws, "C32", datos.get("pol", ""))
    set_valor(ws, "B34", datos.get("pod", ""))

    # -------------------------
    # FECHA
    # -------------------------
    fecha = formatear_fecha(datos.get("fecha_bl", ""))
    celda_fecha = set_valor(ws, "F56", fecha)

    if isinstance(fecha, datetime):
        ws[celda_fecha].number_format = "DD-MM-YYYY"

    # -------------------------
    # TOTALES
    # -------------------------
    set_valor(ws, "D57", datos.get("freight", ""))
    set_valor(ws, "E54", datos.get("bultos", ""))
    set_valor(ws, "G54", datos.get("peso_total", ""))

    # -------------------------
    # ITEMS
    # -------------------------
    items = datos.get("items", [])

    if not items:
        items = [{"specification": ""}]

    for i, item in enumerate(items):
        fila = fila_inicio + i
        spec = item.get("specification", "")

        celda = set_valor(ws, f"D{fila}", spec)
        ws[celda].alignment = Alignment(wrap_text=True)

        auto_alto_fila(ws, fila, spec)

    fila_fin = fila_inicio + len(items)

    # -------------------------
    # CAMIÓN
    # -------------------------
    make = datos.get("make", "")
    tipo = datos.get("type", "")
    vin = datos.get("vin", "")
    transit = (datos.get("transit") or "").strip()

    # 🔥 SHIPPING EN B19
    if not transit:
        set_valor(ws, "B19", "Y/O SDW SHIPPING CHILE SPA")
    else:
        set_valor(ws, "B19", "")

    filas_camion = [
        ("MAKE", make),
        ("TYPE", tipo),
        ("VIN", vin),
        ("TRANSIT", transit),
    ]

    for i, (label, valor) in enumerate(filas_camion):
        fila = fila_fin + 1 + i

        set_valor(ws, f"C{fila}", label)

        celda = set_valor(ws, f"D{fila}", valor)
        ws[celda].alignment = Alignment(wrap_text=True)

        auto_alto_fila(ws, fila, valor)

    # -------------------------
    # GUARDAR
    # -------------------------
    wb.save(ruta_salida)

    # -------------------------
    # 🔥 MARCA DE AGUA FINAL
    # -------------------------
    agregar_marca_agua_pywin32(ruta_salida)