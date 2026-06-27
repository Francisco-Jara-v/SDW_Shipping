from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from openpyxl.utils import coordinate_to_tuple
import re

# -------------------------
# 🔧 UTILIDADES
# -------------------------

def formatear_fecha(fecha_str):
    try:
        return datetime.strptime(fecha_str, "%d.%m.%Y")
    except:
        return fecha_str


def auto_alto_fila(ws, fila, texto, ancho_aprox=45):
    # Si no hay texto o es muy corto, dejamos que Excel use el alto del template
    # No asignamos nada a row_dimensions[fila].height para que no se resetee
    if not texto or len(str(texto)) < ancho_aprox:
        return 

    # Si es largo, calculamos el alto necesario
    lineas_manuales = str(texto).count('\n') + 1
    lineas_por_longitud = (len(str(texto)) // ancho_aprox) + 1
    total_lineas = max(lineas_manuales, lineas_por_longitud)
    
    # Usamos 18-20 para dar margen de seguridad abajo
    ws.row_dimensions[fila].height = total_lineas * 20

def limpiar_tabla(ws, fila_inicio, max_filas=17):
    for fila in range(fila_inicio, fila_inicio + max_filas):
        for col in ["D", "E"]:
            celda = f"{col}{fila}"
            celda_real = get_celda_principal(ws, celda)

            fila_real, _ = coordinate_to_tuple(celda_real)

            # 🔥 solo limpiar si está dentro de la zona
            if fila_real >= fila_inicio:
                ws[celda_real].value = None


# -------------------------
# 🧠 MANEJO DE MERGE (CLAVE)
# -------------------------

def get_celda_principal(ws, celda):
    fila, col = coordinate_to_tuple(celda)

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        if min_row <= fila <= max_row and min_col <= col <= max_col:
            return merged_range.start_cell.coordinate

    return celda


def set_valor(ws, celda, valor):
    """
    Escribe en la celda correcta (merge-safe)
    """
    celda_real = get_celda_principal(ws, celda)
    ws[celda_real] = valor
    return celda_real


def set_wrap(ws, celda):
    """
    Aplica wrap_text correctamente (merge-safe)
    """
    celda_real = get_celda_principal(ws, celda)
    ws[celda_real].alignment = Alignment(wrap_text=True)
    return celda_real


# -------------------------
# 🚀 GENERADOR PRINCIPAL
# -------------------------

def llenar_excel(ruta_plantilla, ruta_salida, datos):
    wb = load_workbook(ruta_plantilla)
    ws = wb.active

    # -------------------------
    # LIMPIAR TABLA
    # -------------------------
    fila_inicio = 42
    limpiar_tabla(ws, fila_inicio)

    # -------------------------
    # DATOS FIJOS
    # -------------------------
    set_valor(ws, "E10", datos.get("bl_completo", ""))

    set_valor(ws, "B19", datos.get("consignee", ""))
    set_valor(ws, "B20", datos.get("nit", ""))


    set_valor(ws, "B26", datos.get("consignee", ""))
    set_valor(ws, "B27", datos.get("nit", ""))

    set_valor(ws, "B36", datos.get("nave", ""))
    set_valor(ws, "C36", datos.get("pol", ""))
    set_valor(ws, "B38", datos.get("pod", ""))

    # -------------------------
    # FECHA
    # -------------------------
    fecha = formatear_fecha(datos.get("fecha_bl", ""))
    celda_fecha = set_valor(ws, "E66", fecha)

    if isinstance(fecha, datetime):
        ws[celda_fecha].number_format = "DD-MM-YYYY"

    # -------------------------
    # TOTALES
    # -------------------------
    
    freight = datos.get("freight")

    celda_freight = set_valor(ws, "B67", freight)

    if freight:
        ws[celda_freight].value = float(freight)
        ws[celda_freight].number_format = '#,##0'
    else:
        ws[celda_freight].value = ""

    ws[celda_freight].alignment = Alignment(horizontal="center")
    
    set_valor(ws, "B71", datos.get("bultos", ""))
    
    peso = datos.get("peso_total", 0)
    celda = set_valor(ws, "G54", peso)

    ws[celda].number_format = 'General'
    ws[celda].value = float(peso)
    ws[celda].number_format = '#,##0'

    # -------------------------
    # ITEMS
    # -------------------------
    items = datos.get("items", [])

    if not items:
        items = [{"specification": ""}]

    for i, item in enumerate(items):
        fila = fila_inicio + i
        spec = item.get("specification", "")
        contenedor = item.get("delivered_container", "")

        texto_item = f"{i + 1}. {spec}" if spec else f"{i + 1}."
        
        celda = set_valor(ws, f"D{fila}", texto_item)
        ws[celda].alignment = Alignment(wrap_text=True, vertical="center")
        
        if re.match(r'^[A-Za-z]{4}\d+', contenedor):
            celda_container = set_valor(ws, f"B{fila}", contenedor)
            ws[celda_container].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        
        auto_alto_fila(ws, fila, texto_item)

    fila_fin = fila_inicio + len(items)

    # -------------------------
    # CAMIÓN (DINÁMICO)
    # -------------------------
    make = datos.get("make", "")
    tipo = datos.get("type", "")
    vin = datos.get("vin", "")
    transit = datos.get("transit", "")
    
    if not transit:
        set_valor(ws, "B21", "Y/O SDW SHIPPING CHILE SPA")
    else:
        set_valor(ws, "B21", "")

    fila_base = fila_fin
    
    # Marca
    if make:
        set_valor(ws, f"C{fila_base}", "Marca")
        celda = set_valor(ws, f"D{fila_base}", make)
        ws[celda].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[fila].height = 15
        fila_base += 1

    # Modelo (solo si existe)
    if tipo:
        set_valor(ws, f"C{fila_base}", "Modelo")
        celda = set_valor(ws, f"D{fila_base}", tipo)
        ws[celda].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[fila].height = 15
        fila_base += 1

    # Chasis
    if vin:
        set_valor(ws, f"C{fila_base}", "Chasis")
        celda = set_valor(ws, f"D{fila_base}", vin)
        ws[celda].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[fila].height = 10
        fila_base += 1

    # Transit (sin label)
    if transit:
        celda = set_valor(ws, f"D{fila_base}", transit)
        ws[celda].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[fila].height = 10

    # -------------------------
    # GUARDAR
    # -------------------------
    wb.save(ruta_salida)