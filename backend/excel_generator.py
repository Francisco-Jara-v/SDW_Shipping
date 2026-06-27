from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import coordinate_to_tuple
from datetime import datetime
import re

# -------------------------
# 🎨 ALINEACIONES
# -------------------------

ALIGN_CENTER = Alignment(
    horizontal="center",
    vertical="center"
)

ALIGN_WRAP = Alignment(
    wrap_text=True,
    vertical="center"
)

ALIGN_CENTER_WRAP = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

ALIGN_LEFT = Alignment(
    horizontal="left",
    vertical="center"
)

# -------------------------
# 📍 COORDENADAS
# -------------------------

CELDAS = {

    "bl": "E10",

    "consignee1": "B19",
    "nit1": "B20",

    "consignee2": "B27",
    "nit2": "B28",

    "shipping": "B21",

    "nave": "B36",
    "pol": "C36",
    "pod": "B38",

    "fecha": "E66",

    "freight": "B67",
    
    "total_bultos": "B71",

}

# -------------------------
# 📦 TABLA
# -------------------------

TABLA = {

    "fila_inicio": 42,

    "container": "A",
    "bulto": "B",
    #"campo": "C",
    "descripcion": "C",
    "peso": "E"

}

def formatear_fecha(fecha_str):
    try:
        return datetime.strptime(fecha_str, "%d.%m.%Y")
    except:
        return fecha_str
    
def auto_alto_fila(ws, fila, texto, ancho_aprox=45):
    # Si no hay texto o es muy corto, dejamos que Excel use el alto del template
    # No asignamos nada a row_dimensions[fila].height para que no se resetee
    if not texto or len(str(texto)) < ancho_aprox:
        return 
    
def limpiar_tabla(ws):

    fila_inicio = TABLA["fila_inicio"]

    for fila in range(fila_inicio, fila_inicio + 17):

        for col in [
            TABLA["container"],
            TABLA["bulto"],
            #TABLA["campo"],
            TABLA["descripcion"],
            TABLA["peso"]
        ]:

            celda = f"{col}{fila}"

            celda_real = get_celda_principal(ws, celda)

            fila_real, _ = coordinate_to_tuple(celda_real)

            if fila_real >= fila_inicio:
                ws[celda_real].value = None
            
# -------------------------
# 🧠 MANEJO DE MERGE (CLAVE)
# -------------------------

def get_celda_principal(ws, celda):
    fila, col = coordinate_to_tuple(celda)

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        if min_row <= fila <= max_row and min_col <= col <= max_col:
            return merged_range.start_cell.coordinate

    return celda

def set_valor(ws, celda, valor):
    """
    Escribe en la celda correcta (merge-safe)
    """
    celda_real = get_celda_principal(ws, celda)
    ws[celda_real] = valor
    return celda_real

# -------------------------
# 📝 DATOS FIJOS
# -------------------------

def escribir_datos_fijos(ws, datos):

    celda = set_valor(ws, CELDAS["bl"], datos.get("bl_completo", ""))
    ws[celda].alignment = ALIGN_LEFT

    # Consignee
    celda = set_valor(ws, CELDAS["consignee1"], datos.get("consignee", ""))
    ws[celda].alignment = ALIGN_LEFT

    celda = set_valor(ws, CELDAS["nit1"], datos.get("nit", ""))
    ws[celda].alignment = ALIGN_LEFT

    celda = set_valor(ws, CELDAS["consignee2"], datos.get("consignee", ""))
    ws[celda].alignment = ALIGN_LEFT

    celda = set_valor(ws, CELDAS["nit2"], datos.get("nit", ""))
    ws[celda].alignment = ALIGN_LEFT

    # Nave
    celda = set_valor(ws, CELDAS["nave"], datos.get("nave", ""))
    ws[celda].alignment = ALIGN_LEFT

    celda = set_valor(ws, CELDAS["pol"], datos.get("pol", ""))
    ws[celda].alignment = ALIGN_LEFT

    celda = set_valor(ws, CELDAS["pod"], datos.get("pod", ""))
    ws[celda].alignment = ALIGN_LEFT

    # Shipping
    transit = datos.get("transit", "")

    if transit:
        celda = set_valor(ws, CELDAS["shipping"], "")
    else:
        celda = set_valor(ws, CELDAS["shipping"], "Y/O SDW SHIPPING CHILE SPA")

    ws[celda].alignment = ALIGN_LEFT

# -------------------------
# 💰 FECHA Y TOTALES
# -------------------------

def escribir_totales(ws, datos):

    # Fecha
    fecha = formatear_fecha(datos.get("fecha_bl", ""))

    celda_fecha = set_valor(ws, CELDAS["fecha"], fecha)

    if isinstance(fecha, datetime):
        ws[celda_fecha].number_format = "DD-MM-YYYY"

    # Freight
    freight = datos.get("freight")

    celda = set_valor(ws, CELDAS["freight"], freight)

    if freight:
        ws[celda].value = float(freight)
        ws[celda].number_format = "#,##0"
    else:
        ws[celda].value = ""

    ws[celda].alignment = ALIGN_CENTER
    
    # Bultos
    celda = set_valor(
        ws,
        CELDAS["total_bultos"],
        datos.get("bultos", 0)
    )

    ws[celda].alignment = ALIGN_CENTER

# -------------------------
# 📦 ITEMS
# -------------------------

def escribir_items(ws, datos):

    fila = TABLA["fila_inicio"]

    items = datos.get("items", [])

    if not items:
        items = [{"specification": ""}]

    peso = datos.get("peso_total", 0)

    for i, item in enumerate(items):

        fila_actual = fila + i

        spec = item.get("specification", "")
        contenedor = item.get("delivered_container", "")

        # -----------------
        # Container
        # -----------------

        if re.match(r'^[A-Za-z]{4}\d+', contenedor):

            celda = set_valor(
                ws,
                f"{TABLA['container']}{fila_actual}",
                contenedor
            )

            ws[celda].alignment = ALIGN_CENTER

        # -----------------
        # Nº Bulto
        # -----------------

        celda = set_valor(
            ws,
            f"{TABLA['bulto']}{fila_actual}",
            i + 1
        )

        ws[celda].alignment = ALIGN_CENTER

        # -----------------
        # Specification
        # -----------------

        celda = set_valor(
            ws,
            f"{TABLA['descripcion']}{fila_actual}",
            spec
        )

        ws[celda].alignment = ALIGN_WRAP

        # -----------------
        # Peso
        # -----------------

        celda = set_valor(
            ws,
            f"{TABLA['peso']}{fila_actual}",
            peso
        )

        ws[celda].number_format = "#,##0"
        ws[celda].alignment = ALIGN_CENTER

        auto_alto_fila(ws, fila_actual, spec)

    return fila + len(items)

# -------------------------
# 🚛 VEHÍCULO
# -------------------------

def escribir_vehiculo(ws, datos, fila):

    make = datos.get("make", "")
    tipo = datos.get("type", "")
    vin = datos.get("vin", "")
    transit = datos.get("transit", "")

    if make:

        celda = set_valor(
            ws,
            f"{TABLA['descripcion']}{fila}",
            f"Marca: {make}"
        )

        ws[celda].alignment = ALIGN_LEFT
        fila += 1

    if tipo:

        celda = set_valor(
            ws,
            f"{TABLA['descripcion']}{fila}",
            f"Modelo: {tipo}"
        )

        ws[celda].alignment = ALIGN_LEFT
        fila += 1

    if vin:

        celda = set_valor(
            ws,
            f"{TABLA['descripcion']}{fila}",
            f"Chasis: {vin}"
        )

        ws[celda].alignment = ALIGN_LEFT
        fila += 1

    if transit:

        celda = set_valor(
            ws,
            f"{TABLA['descripcion']}{fila}",
            f"Transit: {transit}"
        )

        ws[celda].alignment = ALIGN_LEFT

# -------------------------
# 🚀 GENERADOR PRINCIPAL
# -------------------------

def llenar_excel(ruta_plantilla, ruta_salida, datos):

    wb = load_workbook(ruta_plantilla)

    # Recorre todas las hojas del libro
    # (Original y Copia no negociable)
    for ws in wb.worksheets:

        # Limpiar tabla
        limpiar_tabla(ws)

        # Datos del encabezado
        escribir_datos_fijos(ws, datos)

        # Fecha y Freight
        escribir_totales(ws, datos)

        # Items
        fila_siguiente = escribir_items(ws, datos)

        # Datos del vehículo
        escribir_vehiculo(ws, datos, fila_siguiente)

    wb.save(ruta_salida)